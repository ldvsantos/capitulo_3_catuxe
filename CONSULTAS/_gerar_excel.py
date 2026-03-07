import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Estilos ──────────────────────────────────────────────
bold = Font(bold=True, size=11, name="Calibri")
bold_white = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
normal = Font(size=11, name="Calibri")
italic = Font(italic=True, size=11, name="Calibri")
title_font = Font(bold=True, size=14, name="Calibri")
subtitle_font = Font(bold=True, size=12, name="Calibri")
wrap = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
dim_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
question_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def apply_border(ws, row, cols):
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = thin_border

# ── Dados das dimensões (arquitetura V1–V8) ──────────────
dimensions = [
    {
        "code": "V1",
        "name": "Erosão Intergeracional e Migração",
        "definition": (
            "Risco de descontinuidade na transmissão vertical do conhecimento tradicional "
            "entre gerações, agravado por dinâmicas de êxodo rural e migração juvenil. "
            "Caracteriza-se pela redução progressiva do número de jovens que dominam saberes "
            "e práticas agrícolas tradicionais, pela diminuição da frequência de eventos de "
            "ensino-aprendizagem intergeracional (rodas de conversa, acompanhamento na roça, "
            "mutirões com aprendizes) e pela perda de significados rituais, espirituais ou "
            "cosmológicos vinculados às práticas de manejo. Manifesta-se quando o saber "
            "permanece restrito a anciãos sem mecanismos ativos de repasse, quando modalidades "
            "orais de transmissão são substituídas por abandono, quando a dimensão espiritual "
            "da prática se dissipa por desinteresse ou pressão externa, ou quando a saída de "
            "jovens para centros urbanos interrompe a cadeia de sucessão do conhecimento."
        ),
        "items": [
            ("9.1.1", "Qual a proporção de jovens (<35 anos) que dominam esta prática/saber na comunidade?\n(Nenhum / Poucos <25% / Alguns 25–50% / A maioria 50–75% / Quase todos >75%)"),
            ("9.1.2", "Com que frequência ocorrem eventos de ensino/aprendizagem desta prática entre gerações (rodas de conversa, acompanhamento na roça, mutirões com aprendizes)?\n(Nunca / Raramente <1x/ano / Anualmente / Semestralmente / Mensalmente ou mais)"),
            ("9.1.3", "Como este saber é transmitido? [múltipla escolha]\n□ Oralidade (conversa, narrativa, cantos)\n□ Prática acompanhada (aprendiz trabalha junto ao mestre)\n□ Ritual ou cerimônia (bênção de sementes, festas de colheita)\n□ Registro escrito ou audiovisual produzido pela comunidade\n□ Não está sendo transmitido atualmente\n□ Outro"),
            ("9.1.4", "Esta prática possui dimensão espiritual, ritual ou cosmológica (p.ex. bênçãos sobre sementes, plantio associado a ciclos lunares, proibições em datas sagradas)?\n(Não / Sim, secundário / Sim, central)"),
            ("9.1.5", "Na sua avaliação, qual o risco de este saber desaparecer na próxima geração?\n(Muito baixo / Baixo / Moderado / Alto / Muito alto)"),
        ]
    },
    {
        "code": "V2",
        "name": "Complexidade e Singularidade Biocultural",
        "definition": (
            "Riqueza, profundidade e exclusividade geográfica das interações ecológicas "
            "codificadas nos sistemas agrícolas tradicionais. Integra dois componentes "
            "complementares. O primeiro refere-se à diversidade de variedades locais (crioulas) "
            "cultivadas ou manejadas, às interações solo-planta-clima reconhecidas pelo agricultor "
            "segundo taxonomia êmica (local) e à agrobiodiversidade mantida pelo sistema de manejo "
            "tradicional. O segundo componente refere-se à exclusividade geográfica dessas práticas, "
            "variedades e saberes em relação a outras comunidades da região, ou seja, ao grau em que "
            "determinados conhecimentos, cultivares ou técnicas são endêmicos de uma comunidade "
            "específica e dependem de condições ambientais particulares do local (microclima, solo, "
            "hidrologia, vegetação nativa). A singularidade territorial é operacionalizada pela "
            "beta-diversidade de práticas entre comunidades e pela proporção de práticas exclusivas. "
            "Quanto maior a singularidade, maior o risco de perda irreversível caso a comunidade "
            "abandone a prática, pois não há redundância geográfica."
        ),
        "items": [
            ("9.3.1", "Inventário de variedades locais: para cada cultura principal, listar todas as variedades locais (crioulas, tradicionais) cultivadas ou conhecidas pelo informante, usando nomes locais. Registrar: cultura, nome da variedade, origem/história e uso principal."),
            ("9.3.2", "Classificação êmica de solos: registrar como o agricultor classifica os tipos de solo do território, usando a terminologia local (nome local, características segundo o agricultor, culturas associadas). O pesquisador indica correspondência pedológica aproximada."),
            ("9.3.3", "O agricultor segue regras de associação entre tipo de solo, variedade cultivada e época de plantio baseadas em conhecimento tradicional?\n(Não / Sim). Se sim, descrever até 3 regras principais."),
            ("9.4.1", "Até onde o agricultor sabe, esta prática/variedade existe em outras comunidades da região?\n(Sim, é comum / Existe em algumas comunidades / Só existe nesta comunidade / Não sabe)"),
            ("9.4.2", "Esta prática depende de condições ambientais específicas deste território (tipo de solo, microclima, água, vegetação nativa)?\n(Não, pode ser aplicada em qualquer lugar / Parcialmente / Sim, totalmente, só funciona aqui)"),
            ("9.4.3", "Inventário comparativo entre comunidades: para cada prática/saber documentado, indicar presença ou ausência em cada comunidade visitada (preenchido pelo pesquisador após levantamento multi-sítio)."),
        ]
    },
    {
        "code": "V3",
        "name": "Status de Documentação",
        "definition": (
            "Grau de registro formal dos saberes e práticas agrícolas tradicionais da comunidade. "
            "Refere-se à existência de documentação prévia (publicações acadêmicas, relatórios técnicos, "
            "fichas WOCAT, registros audiovisuais, inventários patrimoniais, registros comunitários) e à "
            "completude dessa documentação. Avalia também o risco de perda documental por deterioração de "
            "mídias, falecimento de informantes-chave ou falta de infraestrutura de armazenamento. Um "
            "saber com baixo status de documentação é aquele que existe apenas na memória oral de poucos "
            "indivíduos, sem nenhum tipo de registro formal, tornando-o vulnerável à perda silenciosa e "
            "irreversível. A dimensão V3 funciona como diagnóstico de urgência: quanto menor o status de "
            "documentação, mais prioritária a ação de salvaguarda."
        ),
        "items": [
            ("9.5.1", "Este saber/prática já foi objeto de algum registro formal antes deste levantamento?\n(Não / Sim, parcialmente / Sim, completamente)\nSe sim, tipo: □ Publicação acadêmica □ Relatório técnico □ Ficha WOCAT □ Audiovisual □ Inventário IPHAN □ Registro comunitário □ Outro"),
            ("9.5.2", "Completude documental: para o conjunto de saberes/práticas identificados nesta comunidade, informar:\n(a) nº total identificados\n(b) nº com registro formal\n(c) nº com ficha técnica completa\n(d) nº com registro audiovisual\n(e) nº digitalizados e acessíveis online"),
            ("9.5.3", "Existem acervos orais (gravações, relatos) em risco de perda por deterioração de mídias, falecimento de informantes ou falta de infraestrutura de armazenamento?\n(Não / Sim). Se sim, descrever."),
        ]
    },
    {
        "code": "V4",
        "name": "Vulnerabilidade Jurídica e Fundiária",
        "definition": (
            "Exposição dos saberes e práticas agrícolas tradicionais a apropriação indevida por terceiros, "
            "decorrente da ausência ou insuficiência de instrumentos jurídicos de proteção. Abrange a "
            "situação fundiária da comunidade (certificação Palmares, titulação INCRA, conflitos), a "
            "existência de mecanismos de propriedade intelectual aplicáveis (Indicação Geográfica, marca "
            "coletiva, registro IPHAN, protocolo comunitário sob Nagoya/Kunming-Montreal, registro de "
            "cultivar tradicional, selos e certificações), o grau de conhecimento da comunidade sobre "
            "esses mecanismos e a ocorrência de casos de apropriação indevida (biopirataria, uso comercial "
            "sem consentimento, patenteamento de variedades tradicionais). Uma comunidade com alta "
            "vulnerabilidade jurídica detém saberes valiosos sem nenhuma camada formal de proteção e "
            "desconhece os instrumentos disponíveis para protegê-los."
        ),
        "items": [
            ("9.6.1", "Situação fundiária:\n• Certificação Palmares (Sim/Não/Em processo)\n• Título coletivo INCRA/ITERBA (Sim/Não/Em processo)\n• Conflito fundiário ativo (Sim/Não)\n• Área titulada (ha) / Área reivindicada (ha)"),
            ("9.6.2", "Instrumentos de PI (para cada um, indicar situação: Não / Em processo / Sim):\n• IG (INPI) • Marca Coletiva (INPI) • Patrimônio Imaterial (IPHAN) • Registro estadual/municipal • Protocolo Comunitário • Cultivar tradicional (MAPA/RNC) • Selo/certificação"),
            ("9.6.3", "A comunidade tem conhecimento da existência dos instrumentos legais de proteção?\n(Desconhece totalmente / Conhece vagamente / Conhece mas não sabe acessar / Conhece e já tentou acessar)\nJá recebeu assessoria jurídica? (Não / Sim, especificar instituição)"),
            ("9.6.4", "Há registro ou percepção de que saberes, variedades ou práticas tenham sido apropriados por terceiros sem consentimento ou repartição de benefícios?\n(Não / Sim / Suspeita sem confirmação). Se sim, descrever."),
        ]
    },
    {
        "code": "V5",
        "name": "Organização Social e Governança",
        "definition": (
            "Vitalidade das estruturas comunitárias de governança que sustentam a manutenção e a "
            "transmissão dos saberes agrícolas tradicionais. Refere-se à existência de mestres de saberes "
            "(guardiões do conhecimento reconhecidos pela comunidade), à frequência de eventos coletivos "
            "de prática e transmissão (mutirões, trocas de sementes, feiras de agrobiodiversidade, rodas "
            "de prosa, encontros intercomunitários), à existência de redes de cooperação com outras "
            "comunidades quilombolas ou tradicionais, e à equidade de gênero tanto na detenção do "
            "conhecimento quanto na governança das decisões sobre manejo. Uma comunidade com organização "
            "social fragilizada é aquela onde os saberes estão atomizados em indivíduos isolados, sem "
            "eventos coletivos de prática, sem articulação intercomunitária e com assimetrias de gênero "
            "na governança, reduzindo a resiliência do sistema de conhecimento."
        ),
        "items": [
            ("9.2.1", "Quantas pessoas na comunidade são reconhecidas como guardiãs/mestres deste saber?\n(Nenhuma / 1–2 / 3–5 / 6–10 / >10)\nFaixa etária predominante: □ Jovens <35  □ Meia-idade 35–59  □ Idosos ≥60"),
            ("9.2.2", "Com que frequência ocorrem eventos coletivos relacionados a este saber (mutirões, trocas de sementes, feiras de agrobiodiversidade, rodas de prosa, encontros intercomunitários)?\n(Nunca / Raramente <1x/ano / Anualmente / Semestralmente / Mensalmente ou mais)"),
            ("9.2.3", "A comunidade mantém intercâmbio deste saber com outras comunidades quilombolas ou tradicionais?\n(Não / Sim, ocasionalmente, contato informal / Sim, regularmente, rede organizada)\nSe sim, especificar comunidades ou redes."),
            ("9.2.4", "Quem participa das decisões sobre o manejo desta prática?\n□ Predominantemente mulheres □ Predominantemente homens □ Ambos equilibrado □ Conselhos de anciãos □ Assembleia comunitária □ Outro\nQuem detém o conhecimento principal?\n□ Predominantemente mulheres □ Predominantemente homens □ Ambos"),
        ]
    },
    {
        "code": "V6",
        "name": "Vitalidade Linguística",
        "definition": (
            "Grau de manutenção e transmissão da língua ou dialeto tradicional da comunidade, "
            "com ênfase na retenção do léxico etnotaxonômico vinculado às práticas de manejo "
            "agrícola e ambiental. Refere-se à proficiência geracional na língua ou falar local "
            "(incluindo variantes do português quilombola, léxico de origem africana, vocabulário "
            "ritual), à frequência de uso dessa língua em contextos de prática agrícola e "
            "transmissão de saberes, e à preservação de termos técnicos tradicionais para solos, "
            "plantas, animais, fenômenos climáticos e ciclos produtivos. A erosão linguística "
            "compromete a transmissão de saberes codificados em categorias êmicas que não possuem "
            "equivalente direto na língua dominante. A dimensão V6 é parcialmente capturada pelas "
            "seções WOCAT originais (§5.6, §6.1), mas requer módulo focal complementar para aferir "
            "a profundidade do uso linguístico-cognitivo."
        ),
        "items": [
            ("9.7.1", "A comunidade possui língua, dialeto ou falar tradicional distinto do português padrão (p.ex. léxico de origem africana, vocabulário ritual, variantes fonológicas próprias)?\n(Não / Sim, resíduo lexical apenas / Sim, falar ativo em parte da comunidade / Sim, falar ativo na maioria)"),
            ("9.7.2", "Qual a proporção de jovens (<35 anos) que compreendem e utilizam o vocabulário tradicional associado às práticas de manejo (nomes de solos, plantas, épocas de plantio, técnicas)?\n(Nenhum / Poucos <25% / Alguns 25–50% / A maioria 50–75% / Quase todos >75%)"),
            ("9.7.3", "Durante as atividades agrícolas e de manejo, qual língua ou vocabulário predomina na comunicação entre gerações?\n(Apenas português padrão / Maioria português com termos tradicionais esporádicos / Mistura equilibrada / Predominância do falar tradicional)"),
        ]
    },
    {
        "code": "V7",
        "name": "Integração ao Mercado",
        "definition": (
            "Grau de inserção comercial dos produtos e práticas agrícolas da comunidade e sua "
            "influência sobre a manutenção ou substituição de repertórios bioculturais locais. "
            "Refere-se à proporção de produção destinada à venda versus autoconsumo, à adoção "
            "de cultivares comerciais em substituição a variedades crioulas, à dependência de "
            "insumos externos e à monetização de práticas antes não comerciais. Comunidades com "
            "alta integração ao mercado podem abandonar variedades e técnicas tradicionais de "
            "menor rendimento comercial, reduzindo a agrobiodiversidade funcional do sistema. "
            "A dimensão V7 é majoritariamente capturada pelas seções originais do WOCAT "
            "(§5.6, §5.9, §6.1, §6.5) e não requer bloco suplementar extenso, mas itens de "
            "verificação permitem ao juiz avaliar a adequação dessa cobertura."
        ),
        "items": [
            ("9.8.1", "Qual a proporção aproximada da produção agrícola da família/comunidade destinada à venda (mercado, feira, atravessador)?\n(Nenhuma, 100% autoconsumo / Pouca <25% / Moderada 25–50% / Maioria 50–75% / Quase toda >75%)"),
            ("9.8.2", "Nos últimos 10 anos, houve substituição de variedades crioulas por cultivares comerciais?\n(Não / Sim, parcialmente / Sim, predominantemente)\nSe sim, indicar culturas afetadas e motivo principal."),
        ]
    },
    {
        "code": "V8",
        "name": "Exposição Climática",
        "definition": (
            "Sensibilidade do sistema agrícola tradicional a variações climáticas e eventos "
            "extremos (seca prolongada, ondas de calor, chuvas concentradas, geadas atípicas). "
            "Refere-se à percepção do agricultor sobre mudanças na regularidade das chuvas e "
            "temperaturas, à frequência e severidade de perdas produtivas atribuídas a eventos "
            "climáticos e à existência de estratégias adaptativas tradicionais (mudança de época "
            "de plantio, seleção de variedades tolerantes, diversificação de cultivos). A dimensão "
            "V8 é majoritariamente operacionalizada pelas seções climáticas originais do WOCAT "
            "(§5.1, §5.4, §6.3) e não requer bloco suplementar extenso, mas itens de verificação "
            "permitem ao juiz avaliar a adequação dessa cobertura."
        ),
        "items": [
            ("9.9.1", "Na percepção do agricultor, o regime de chuvas e as temperaturas mudaram nos últimos 10–20 anos?\n(Não percebeu mudança / Sim, mudança leve / Sim, mudança forte)\nSe sim, descrever as principais mudanças percebidas."),
            ("9.9.2", "Nos últimos 5 anos, quantas vezes a produção foi severamente afetada por eventos climáticos (seca, chuva excessiva, calor extremo, geada)?\n(Nenhuma / 1–2 vezes / 3–4 vezes / 5 ou mais vezes)\nEvento mais grave: _______________"),
        ]
    },
]

# ═══════════════════════════════════════════════════════════
# ABA 1: Juízes Experts
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Juízes Experts"

# Larguras
ws.column_dimensions["A"].width = 45
ws.column_dimensions["B"].width = 65
ws.column_dimensions["C"].width = 14
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 35

# ── Cabeçalho de instruções ──
row = 1
ws.cell(row=row, column=1, value="Ficha de Validade de Conteúdo, Juízes Experts").font = title_font
ws.merge_cells("A1:F1")
row = 2
ws.cell(row=row, column=1, value="Questionário WOCAT-SLM-QBR: Itens Suplementares para Avaliação de Vulnerabilidade Biocultural (V1–V8)").font = subtitle_font
ws.merge_cells("A2:F2")

row = 4
ws.cell(row=row, column=1, value="Instruções").font = bold
row = 5
instr = (
    'Os itens abaixo são candidatos ao questionário WOCAT-SLM-QBR (versão adaptada para comunidades '
    'quilombolas brasileiras). Cada item foi elaborado para operacionalizar uma dimensão de vulnerabilidade '
    'biocultural. A primeira coluna apresenta a definição constitutiva dessa dimensão. A segunda coluna '
    'apresenta os itens pensados para esta dimensão. Por favor, avalie a clareza da linguagem de cada item, '
    'o quanto o item é pertinente para a dimensão e o quanto é relevante para o instrumento. Por fim, '
    'indique se há necessidades de modificação do item.'
)
ws.cell(row=row, column=1, value=instr).font = normal
ws.cell(row=row, column=1).alignment = wrap
ws.merge_cells("A5:F5")
ws.row_dimensions[5].height = 60

row = 7
ws.cell(row=row, column=1, value="1) Clareza: Avalie o quão clara e compreensível está a formulação do item.").font = italic
ws.merge_cells("A7:F7")
row = 8
ws.cell(row=row, column=1, value="2) Pertinência: Se o item representa o construto/fenômeno que a dimensão pretende medir.").font = italic
ws.merge_cells("A8:F8")
row = 9
ws.cell(row=row, column=1, value="3) Relevância: Se o item é relevante para o instrumento como um todo.").font = italic
ws.merge_cells("A9:F9")

row = 11
ws.cell(row=row, column=1, value="Escala: 1 = Inadequado | 2 = Pouco adequado | 3 = Razoável | 4 = Adequado | 5 = Muito adequado").font = bold
ws.merge_cells("A11:F11")

row = 13

# ── Iterar dimensões ──
for dim in dimensions:
    # Cabeçalho da dimensão
    ws.cell(row=row, column=1, value="Definição Constitutiva").font = bold_white
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=1).alignment = wrap_center
    ws.cell(row=row, column=2, value="Item").font = bold_white
    ws.cell(row=row, column=2).fill = header_fill
    ws.cell(row=row, column=2).alignment = wrap_center
    ws.cell(row=row, column=3, value="Clareza\n(1-5)").font = bold_white
    ws.cell(row=row, column=3).fill = header_fill
    ws.cell(row=row, column=3).alignment = wrap_center
    ws.cell(row=row, column=4, value="Pertinência\n(1-5)").font = bold_white
    ws.cell(row=row, column=4).fill = header_fill
    ws.cell(row=row, column=4).alignment = wrap_center
    ws.cell(row=row, column=5, value="Relevância\n(1-5)").font = bold_white
    ws.cell(row=row, column=5).fill = header_fill
    ws.cell(row=row, column=5).alignment = wrap_center
    ws.cell(row=row, column=6, value="Sugestões de alteração").font = bold_white
    ws.cell(row=row, column=6).fill = header_fill
    ws.cell(row=row, column=6).alignment = wrap_center
    apply_border(ws, row, 6)
    row += 1

    # Nome da dimensão
    dim_label = f"{dim['code']} | {dim['name']}"
    ws.cell(row=row, column=1, value=dim_label).font = bold
    ws.cell(row=row, column=1).fill = dim_fill
    ws.cell(row=row, column=1).alignment = wrap
    for c in range(2, 7):
        ws.cell(row=row, column=c).fill = dim_fill
    apply_border(ws, row, 6)
    row += 1

    # Definição + itens
    first_item_row = row
    for i, (item_num, item_text) in enumerate(dim["items"]):
        cell_item = ws.cell(row=row, column=2, value=f"[{item_num}] {item_text}")
        cell_item.font = normal
        cell_item.alignment = wrap
        ws.cell(row=row, column=3).alignment = wrap_center
        ws.cell(row=row, column=4).alignment = wrap_center
        ws.cell(row=row, column=5).alignment = wrap_center
        ws.cell(row=row, column=6).alignment = wrap
        apply_border(ws, row, 6)

        # Altura adequada
        n_lines = item_text.count("\n") + 1
        ws.row_dimensions[row].height = max(30, n_lines * 16)

        row += 1

    # Mesclar coluna A (definição) para os itens
    last_item_row = row - 1
    if last_item_row > first_item_row:
        ws.merge_cells(start_row=first_item_row, start_column=1,
                       end_row=last_item_row, end_column=1)
    ws.cell(row=first_item_row, column=1, value=dim["definition"]).font = normal
    ws.cell(row=first_item_row, column=1).alignment = wrap

    # Pergunta de lacuna
    row += 1
    ws.cell(row=row, column=1, value="Algum componente importante do construto não foi abordado nos itens? Se sim, qual?").font = italic
    ws.cell(row=row, column=1).fill = question_fill
    ws.cell(row=row, column=1).alignment = wrap
    ws.cell(row=row, column=2).fill = question_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    apply_border(ws, row, 6)
    ws.row_dimensions[row].height = 35
    row += 1

    ws.cell(row=row, column=1, value="Sugestão de item adicional?").font = italic
    ws.cell(row=row, column=1).fill = question_fill
    ws.cell(row=row, column=1).alignment = wrap
    ws.cell(row=row, column=2).fill = question_fill
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    apply_border(ws, row, 6)
    ws.row_dimensions[row].height = 35
    row += 2

# ── Avaliação geral ──
row += 1
ws.cell(row=row, column=1, value="Avaliação Geral do Instrumento").font = title_font
ws.cell(row=row, column=1).fill = header_fill
ws.cell(row=row, column=1).font = bold_white
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1

general_qs = [
    "Há alguma dimensão de vulnerabilidade biocultural relevante para SSAT quilombolas que não foi contemplada pelas oito dimensões (V1–V8)? Se sim, qual?",
    "Os itens, no conjunto, são suficientes para operacionalizar as dimensões propostas?",
    "A linguagem dos itens é adequada para aplicação em contexto de comunidades quilombolas rurais do semiárido?",
    "Comentários gerais ou observações adicionais:",
]
for q in general_qs:
    ws.cell(row=row, column=1, value=q).font = normal
    ws.cell(row=row, column=1).alignment = wrap
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    apply_border(ws, row, 6)
    ws.row_dimensions[row].height = 40
    row += 1

# ── Dados do avaliador ──
row += 2
ws.cell(row=row, column=1, value="Dados do Avaliador").font = title_font
ws.cell(row=row, column=1).fill = header_fill
ws.cell(row=row, column=1).font = bold_white
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1

fields = [
    "Nome completo",
    "Formação / área de especialidade",
    "Instituição",
    "Experiência com comunidades tradicionais (anos)",
    "Data da avaliação",
]
for f in fields:
    ws.cell(row=row, column=1, value=f).font = bold
    ws.cell(row=row, column=1).alignment = wrap
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    apply_border(ws, row, 6)
    ws.row_dimensions[row].height = 25
    row += 1

# ── Salvar ──
out_path = r"C:\Users\vidal\OneDrive\Documentos\13 - CLONEGIT\artigo_3_catuxe\CONSULTAS\Ficha_Validade_Conteudo_Juizes_V1V8.xlsx"
wb.save(out_path)
print(f"Arquivo salvo em: {out_path}")
